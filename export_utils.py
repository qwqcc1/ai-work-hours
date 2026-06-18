import io
import re
import os
from datetime import date
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from crud import get_work_entries


TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template.xlsx")


def export_to_excel(db: Session, start_date: date = None, end_date: date = None,
                    project_id: int = None, user_id: int = None,
                    limit: int = 10000) -> io.BytesIO:
    entries = get_work_entries(db, user_id=user_id, project_id=project_id,
                               start_date=start_date, end_date=end_date, limit=limit)

    wb = Workbook()
    ws = wb.active
    ws.title = "工时数据"

    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    cell_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["序号", "人员", "项目名称", "项目分类", "项目番号", "完成情况",
               "地区", "工时类型", "工时(h)", "填报日期", "设备来源", "工作描述", "录入时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, entry in enumerate(entries, 2):
        values = [
            row_idx - 1,
            entry.user.real_name if entry.user else "",
            entry.project.name if entry.project else "",
            entry.project.category if entry.project else "",
            entry.project_number or "",
            entry.completion_status or "",
            entry.location or "",
            entry.work_type,
            entry.hours,
            str(entry.entry_date),
            entry.device_info or "PC",
            entry.description,
            str(entry.created_at)
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = cell_alignment
            cell.border = thin_border

    widths = [6, 12, 20, 12, 12, 10, 12, 14, 8, 14, 12, 30, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # 汇总工作表
    ws2 = wb.create_sheet("汇总统计")
    total_hours = sum(e.hours for e in entries)
    ws2.cell(row=1, column=1, value="汇总信息").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="总工时(h)").font = Font(bold=True)
    ws2.cell(row=3, column=2, value=round(total_hours, 1))
    ws2.cell(row=4, column=1, value="记录数").font = Font(bold=True)
    ws2.cell(row=4, column=2, value=len(entries))
    ws2.cell(row=5, column=1, value="导出时间").font = Font(bold=True)
    ws2.cell(row=5, column=2, value=str(date.today()))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_month_column_map(ws, month: int) -> dict:
    """Map day number -> column index for a given month in the worksheet."""
    current_month = None
    day_map = {}
    for col in range(1, ws.max_column + 1):
        v1 = ws.cell(row=1, column=col).value
        v3 = ws.cell(row=3, column=col).value
        if v1 and '月份工时统计' in str(v1):
            m = re.search(r'(\d+)\s*月份', str(v1))
            if m:
                current_month = int(m.group(1))
        elif current_month and current_month > month:
            break
        if current_month == month and v3 and str(v3).isdigit():
            day_map[int(v3)] = col
    return day_map


def _unmerge_rows(ws, start_row: int, end_row: int):
    """Unmerge any merged cell ranges that overlap with the given row range."""
    to_unmerge = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= end_row and mc.max_row >= start_row:
            to_unmerge.append(str(mc))
    for ref in to_unmerge:
        ws.unmerge_cells(ref)


def _find_baogong_row(ws) -> int:
    """Find the row index of the '保有工时' summary row."""
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=6).value == '保有工时':
            return row
    return ws.max_row


def _save_row_data(ws, row: int, max_col: int) -> dict:
    """Save all cell values from a row."""
    data = {}
    for col in range(1, max_col + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            data[col] = val
    return data


def _restore_row_with_fixed_refs(ws, target_row: int, data: dict, old_row: int,
                                 formulas_only: bool = False):
    """Restore saved cell values to a row, fixing formula row references.
    If formulas_only is True, only restore formula-type values."""
    for col, val in data.items():
        cell = ws.cell(row=target_row, column=col)
        if isinstance(val, str) and val.startswith('='):
            fixed = re.sub(r'(?<=[A-Z])' + str(old_row) + r'(?![0-9])', str(target_row), val)
            cell.value = fixed
        elif not formulas_only:
            cell.value = val


def fill_monthly_template(db: Session, month: int, year: int = 2026) -> io.BytesIO:
    """
    Fill the annual timesheet template with work hours for a specific month.
    Each employee gets a row, hours filled into the correct day columns.
    The '保有工时' summary row is preserved and always kept at the bottom.
    If there aren't enough rows for employees, rows are auto-inserted.
    """
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Get column mapping for this month
    day_map = _build_month_column_map(ws, month)
    if not day_map:
        raise ValueError(f"Could not find columns for month {month} in template")

    # Build date range for the month
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)

    # Get all users who have work entries in this month
    entries = get_work_entries(db, start_date=start_date, end_date=end_date, limit=10000)

    # Group entries by (user, project_number) — one row per user per project
    grouped = {}
    for e in entries:
        uid = e.user_id
        pn = e.project_number or ''
        key = (uid, pn)
        if key not in grouped:
            username = e.user.username if e.user else f'User {uid}'
            grouped[key] = {
                'name': username,
                'project_number': pn,
                'entries': {},
                'completion_statuses': {}
            }
        day = e.entry_date.day
        grouped[key]['entries'][day] = grouped[key]['entries'].get(day, 0) + e.hours
        cs = e.completion_status or ''
        if cs:
            grouped[key]['completion_statuses'][cs] = grouped[key]['completion_statuses'].get(cs, 0) + 1

    # Save template data before any modifications
    baogong_orig = _find_baogong_row(ws)
    baogong_data = _save_row_data(ws, baogong_orig, ws.max_column)
    # Save template employee row formulas (row 4) for weekly/monthly subtotals
    template_formulas = _save_row_data(ws, 4, ws.max_column)

    first_employee_row = 4
    template_slots = baogong_orig - first_employee_row

    row_list = list(grouped.values())
    needed_rows = len(row_list) if row_list else 1

    # Unmerge everything in the data area (including 保有工时 row)
    _unmerge_rows(ws, first_employee_row, baogong_orig)

    # Delete ALL old rows (employee area + 保有工时)
    total_to_delete = baogong_orig - first_employee_row + 1
    ws.delete_rows(first_employee_row, total_to_delete)

    # Insert fresh employee rows, then the 保有工时 row at the end
    ws.insert_rows(first_employee_row, needed_rows + 1)

    # Fill employee rows (one row per user per project)
    if row_list:
        for i, item in enumerate(row_list):
            row = first_employee_row + i
            ws.cell(row=row, column=1, value='项目运维部')
            ws.cell(row=row, column=2, value=item['name'])
            ws.cell(row=row, column=3, value=item['project_number'])
            cs = max(item['completion_statuses'], key=item['completion_statuses'].get) if item['completion_statuses'] else ''
            ws.cell(row=row, column=5, value=cs)
            ws.cell(row=row, column=6, value='实际工时')

            for day_num, hours in item['entries'].items():
                col = day_map.get(day_num)
                if col:
                    ws.cell(row=row, column=col, value=hours)

            # Apply template formulas (weekly/monthly subtotals) with corrected row refs
            _restore_row_with_fixed_refs(ws, row, template_formulas, 4, formulas_only=True)
    else:
        ws.cell(row=first_employee_row, column=1, value='项目运维部')
        ws.cell(row=first_employee_row, column=6, value='实际工时')
        _restore_row_with_fixed_refs(ws, first_employee_row, template_formulas, 4, formulas_only=True)

    # Restore 保有工时 row at the bottom (last employee row + 1)
    new_baogong = first_employee_row + needed_rows
    _restore_row_with_fixed_refs(ws, new_baogong, baogong_data, baogong_orig)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
